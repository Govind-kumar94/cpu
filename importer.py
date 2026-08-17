import csv
import json
import os

from openpyxl import load_workbook

from process import Process


class Importer:

    # -------------------------------------------------
    # SAFE INTEGER CONVERSION
    # -------------------------------------------------

    @staticmethod
    def to_int(value, default=0):

        try:
            return int(value)

        except (TypeError, ValueError):
            return default

    # -------------------------------------------------
    # SAFE STRING
    # -------------------------------------------------

    @staticmethod
    def to_str(value, default=""):

        if value is None:
            return default

        return str(value)

    # -------------------------------------------------
    # IMPORT CSV
    # -------------------------------------------------

    @staticmethod
    def import_csv(filename):

        processes = []

        try:

            with open(
                filename,
                "r",
                encoding="utf-8"
            ) as file:

                reader = csv.DictReader(file)

                for row in reader:

                    if not row:
                        continue

                    pid = row.get("PID", "")

                    if pid == "":
                        continue

                    process = Process(

                        pid=Importer.to_str(pid),

                        arrival_time=Importer.to_int(
                            row.get("Arrival", 0)
                        ),

                        burst_time=Importer.to_int(
                            row.get("Burst", 1),
                            1
                        ),

                        priority=Importer.to_int(
                            row.get("Priority", 1),
                            1
                        )

                    )

                    processes.append(process)

            return processes

        except Exception as e:

            print("CSV Import Error :", e)

            return []

    # -------------------------------------------------
    # IMPORT EXCEL
    # -------------------------------------------------

    @staticmethod
    def import_excel(filename):

        processes = []

        try:

            workbook = load_workbook(filename)

            sheet = workbook.active

            for row in sheet.iter_rows(
                min_row=2,
                values_only=True
            ):

                if row is None:
                    continue

                if row[0] is None:
                    continue

                process = Process(

                    pid=Importer.to_str(row[0]),

                    arrival_time=Importer.to_int(row[1]),

                    burst_time=Importer.to_int(row[2], 1),

                    priority=Importer.to_int(row[3], 1)

                )

                processes.append(process)

            return processes

        except Exception as e:

            print("Excel Import Error :", e)

            return []
        # -------------------------------------------------
    # IMPORT JSON
    # -------------------------------------------------

    @staticmethod
    def import_json(filename):

        processes = []

        try:

            with open(
                filename,
                "r",
                encoding="utf-8"
            ) as file:

                data = json.load(file)

            if not isinstance(data, list):
                return []

            for row in data:

                if not isinstance(row, dict):
                    continue

                pid = row.get("pid", "")

                if pid == "":
                    continue

                process = Process(

                    pid=Importer.to_str(pid),

                    arrival_time=Importer.to_int(
                        row.get("arrival", 0)
                    ),

                    burst_time=Importer.to_int(
                        row.get("burst", 1),
                        1
                    ),

                    priority=Importer.to_int(
                        row.get("priority", 1),
                        1
                    )

                )

                processes.append(process)

            return processes

        except Exception as e:

            print("JSON Import Error :", e)

            return []

    # -------------------------------------------------
    # AUTO IMPORT
    # -------------------------------------------------

    @staticmethod
    def load(filename):

        if not Importer.exists(filename):

            raise FileNotFoundError(
                f"File not found: {filename}"
            )

        extension = os.path.splitext(filename)[1].lower()

        if extension == ".csv":

            return Importer.import_csv(filename)

        elif extension == ".xlsx":

            return Importer.import_excel(filename)

        elif extension == ".json":

            return Importer.import_json(filename)

        else:

            raise ValueError(
                f"Unsupported file format: {extension}"
            )

    # -------------------------------------------------
    # FILE EXISTS
    # -------------------------------------------------

    @staticmethod
    def exists(filename):

        return os.path.isfile(filename)

    # -------------------------------------------------
    # VALIDATE PROCESS LIST
    # -------------------------------------------------

    @staticmethod
    def validate(processes):

        valid = []

        for process in processes:

            if process.burst_time <= 0:
                continue

            if process.arrival_time < 0:
                continue

            if process.priority <= 0:
                process.priority = 1

            valid.append(process)

        return valid

    # -------------------------------------------------
    # LOAD + VALIDATE
    # -------------------------------------------------

    @staticmethod
    def load_and_validate(filename):

        processes = Importer.load(filename)

        return Importer.validate(processes)